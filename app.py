# app.py (Fully Documented)
"""
This module implements a Flask web application for data visualization.

It provides functionalities for uploading data files, performing data
quality checks, applying data fixes, generating plots, and interacting
with a language model for graph interpretation.  It supports both Gemini and
Claude APIs.
"""

import os
import logging
import traceback
from flask import Flask, render_template, request, jsonify, send_from_directory
from werkzeug.utils import secure_filename
import pandas as pd
import base64
import magic
from openpyxl import load_workbook

# Import functions from utils.py
from utils import (
    analyze_data,
    apply_fixes_to_data,
    generate_graph_interpretation_gemini,
    handle_graph_communication_gemini,
    get_plot_suggestion_from_gemini,
    create_dataset_summary,
    generate_graph_interpretation_claude,  # Import Claude functions
    handle_graph_communication_claude,
    get_plot_suggestion_from_claude,
)

# --- Configuration ---
app = Flask(__name__)
"""The Flask application instance."""

app.secret_key = os.urandom(24)
UPLOAD_FOLDER = "uploads"
ALLOWED_MIME_TYPES = {
    "csv": ["text/csv", "text/plain", "application/csv", "application/vnd.ms-excel"],
    "xlsx": ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/zip"],
    "xls": ["application/vnd.ms-excel", "application/octet-stream"],
    "json": ["application/json", "text/plain"],
    "txt": ["text/plain"],
    "tsv": ["text/tab-separated-values", "text/plain"],
    "parquet": ["application/octet-stream", "application/x-parquet"],
    "feather": ["application/octet-stream", "application/x-feather"],
    "orc": ["application/octet-stream", "application/x-orc"],
    "xml": ["application/xml", "text/xml"],
    "html": ["text/html"],
    "hdf5": ["application/x-hdf5", "application/octet-stream"],
    "sql": ["application/sql", "application/x-sql", "application/vnd.sqlite3", "text/plain"],
}
"""
A dictionary mapping allowed file extensions to a list of their acceptable MIME types.
This allows for flexibility in MIME type detection, as different systems and file
contents might lead to variations in the detected MIME type.
"""

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 64 * 1024 * 1024  # 64 MB

# --- Logging Setup ---
logging.basicConfig(filename="app.log", level=logging.DEBUG,
                    format="%(asctime)s - %(levelname)s - %(module)s - %(funcName)s - %(message)s")

# --- Helper Functions ---

def validate_file_content(filepath):
    """Validates the content of an uploaded file using libmagic.

    This function checks if the detected MIME type of the file is among the
    allowed MIME types for its extension.  It uses a dictionary of lists
    to account for variations in MIME type detection.

    Args:
        filepath (str): The path to the file.

    Returns:
        bool: True if the file content is valid, False otherwise.
    """
    print(f"validate_file_content - START: filepath={filepath}")  # Debug print
    try:
        mime = magic.Magic(mime=True)
        detected_type = mime.from_file(filepath)
    except Exception as e:
        logging.exception(f"Error in validate_file_content during MIME detection: {e}")
        print(f"validate_file_content - Error during MIME detection: {e}")
        return False  # Consider any error during MIME detection as invalid

    logging.debug(f"Detected MIME type for {filepath}: {detected_type}")
    print(f"validate_file_content - filepath: {filepath}, detected_type: {detected_type}")  # Debug print

    filename = os.path.basename(filepath)
    filename = secure_filename(filename)
    file_extension = filename.rsplit(".", 1)[1].lower()
    print(f"validate_file_content - filename: {filename}, file_extension: {file_extension}") # Debug Print

    # Check if the extension is allowed AND if the detected MIME type is in the allowed list
    if file_extension in ALLOWED_MIME_TYPES:
        is_allowed = detected_type in ALLOWED_MIME_TYPES[file_extension]
        print(f"validate_file_content - filename: {filename}, extension: {file_extension}, is_allowed: {is_allowed}")  # Debug print
        print(f"validate_file_content - END (allowed): is_allowed={is_allowed}")  # Debug print
        return is_allowed

    logging.debug(f"File type not allowed: {detected_type}")
    print(f"validate_file_content - File type not allowed: {detected_type}")  # Debug print
    print(f"validate_file_content - END (not allowed): detected_type={detected_type}")  # Debug print
    return False

def load_data(filepath, filename):
    """Loads data from a file into a pandas DataFrame.

    This function supports various file formats including CSV, Excel, JSON,
    TSV, Parquet, Feather, ORC, XML, HTML, and HDF5.  It handles potential
    errors like UnicodeDecodeError and provides informative logging.

    Args:
        filepath (str): The path to the file.
        filename (str): The name of the file.

    Returns:
        pandas.DataFrame: The loaded DataFrame, or None if an error occurred.
    """
    print(f"load_data - START: filepath={filepath}, filename={filename}")  # Debug print
    try:
        # Sanitize filename here too
        filename = secure_filename(filename)
        file_extension = filename.rsplit(".", 1)[1].lower()
        logging.info(f"Loading data from {filepath} (extension: {file_extension})")
        print(f"load_data - file_extension: {file_extension}")  # Debug print

        if file_extension == "csv":
            try:
                df = pd.read_csv(filepath, low_memory=False)
            except UnicodeDecodeError:
                logging.warning(f"UnicodeDecodeError reading {filepath}, trying latin1 encoding")
                print(f"load_data - UnicodeDecodeError, trying latin1")  # Debug print
                df = pd.read_csv(filepath, encoding='latin1', low_memory=False)
        elif file_extension in ("xls", "xlsx"):
            try:
                workbook = load_workbook(filename=filepath, read_only=True)
                sheet_name = workbook.sheetnames[0]
                df = pd.read_excel(filepath, sheet_name=sheet_name, engine='openpyxl')
            except Exception as e:
                logging.warning(f"Error loading Excel with openpyxl, trying default engine: {e}")
                print(f"load_data - Error with openpyxl, trying default: {e}")  # Debug print
                df = pd.read_excel(filepath)  # Fallback to default engine
        elif file_extension == "json":
            try:
                df = pd.read_json(filepath, orient="records")
            except ValueError:
                logging.warning(f"ValueError reading JSON with orient=records, trying default")
                print(f"load_data - ValueError with orient=records, trying default")  # Debug print
                df = pd.read_json(filepath)
        elif file_extension == "txt":
            print(f"load_data - txt file detected, returning None")  # Debug print
            return None
        elif file_extension == "tsv":
            df = pd.read_csv(filepath, sep="\t", low_memory=False)
        elif file_extension == "parquet":
            df = pd.read_parquet(filepath)
        elif file_extension == "feather":
            df = pd.read_feather(filepath)
        elif file_extension == "orc":
            df = pd.read_orc(filepath)
        elif file_extension == "xml":
            df = pd.read_xml(filepath)
        elif file_extension == "html":
            df = pd.read_html(filepath)[0]
        elif file_extension == "hdf5":
            with pd.HDFStore(filepath, "r") as store:
                keys = store.keys()
                print(f"load_data - HDF5 keys: {keys}")  # Debug print
                if keys:
                    df = store[keys[0]]
                else:
                    logging.error(f"No datasets found in HDF5 file: {filepath}")
                    print(f"load_data - No datasets in HDF5: {filepath}")  # Debug print
                    raise ValueError("No datasets in HDF5 file.")
        elif file_extension == "sql":
            print(f"load_data - SQL support not implemented.")  # Debug print
            raise NotImplementedError("SQL support not implemented.")
        else:
            logging.error(f"Unsupported file format: {file_extension}")
            print(f"load_data - Unsupported file format: {file_extension}")  # Debug print
            return None  # Return None for unsupported formats

        logging.debug(f"Data loaded successfully. Shape: {df.shape}, Data Types: {df.dtypes}")
        print(f"load_data - Data loaded. Shape: {df.shape}, Data Types: {df.dtypes}")  # Debug print
        print(f"load_data - END (success): df.shape={df.shape}")  # Debug print
        return df

    except Exception as e:
        logging.error(f"Error loading data from {filename}: {e}", exc_info=True)
        print(f"load_data - ERROR: {e}")  # Debug print
        print(f"load_data - END (error): filename={filename}, error={e}")  # Debug print
        return None

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# --- Routes ---

@app.route('/welcome')
def welcome():
    """Renders the welcome page (home.html)."""
    print("welcome - START")  # Debug print
    print("welcome - END")  # Debug print
    return render_template('home.html')

@app.route("/")
def index():
    """Renders the main index page (index.html)."""
    print("index - START") # Debug print
    print("index - END") # Debug print
    return render_template("index.html")

@app.route("/data")
def data():
    """Renders the data information page (data.html)."""
    print("data - START") # Debug print
    print("data - END") # Debug print
    return render_template("data.html")

@app.route("/check")
def check():
    """Renders the data checking page (check.html)."""
    print("check - START") # Debug print
    print("check - END") # Debug print
    return render_template("check.html")

@app.route("/fix")
def fix():
    """Renders the data fixing page (fix.html)."""
    print("fix - START") # Debug print
    print("fix - END") # Debug print
    return render_template("fix.html")

@app.route("/visualisation")
def visualisation():
    """Renders the visualization page (visualisation.html)."""
    print("visualisation - START") # Debug print
    print("visualisation - END") # Debug print
    return render_template("visualisation.html")

@app.route("/upload", methods=["POST"])
def upload_file():
    """Handles file uploads.

    This route accepts a file upload via a POST request.  It checks if a file
    was provided, validates its content using `validate_file_content`, and
    saves it to the `UPLOAD_FOLDER`.  It returns a JSON response indicating
    success or failure.

    Returns:
        flask.Response: A JSON response with a message and status code.
    """
    print("upload_file - START")
    if "file" not in request.files:
        print("upload_file - No file part")
        return jsonify({"error": "No file part"}), 400
    file = request.files["file"]
    if file.filename == "":
        print("upload_file - No selected file")
        return jsonify({"error": "No selected file"}), 400

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)

    # Use a try-except block to handle potential file system errors
    try:
        file.save(filepath)
        print(f"upload_file - filename: {filename}, filepath: {filepath}")

        if validate_file_content(filepath):
            logging.info(f"File uploaded and validated successfully: {filepath}")
            print(f"upload_file - File uploaded and validated successfully: {filepath}")
            print("upload_file - END (success)")
            return jsonify({"message": "File uploaded successfully", "filename": filename}), 200
        else:
            # File is invalid, attempt to remove it
            try:
                os.remove(filepath)
                logging.warning(f"File removed due to invalid content: {filepath}")
                print(f"upload_file - File removed: invalid content: {filepath}")

            except OSError as e:
                logging.exception(f"Error removing invalid file {filepath}: {e}")
                print(f"upload_file - Error removing invalid file: {e}")
                # Even if file removal fails, still report invalid content

            print("upload_file - END (failure)")
            return jsonify({"error": "Invalid file content"}), 400

    except OSError as e:  # Catch file system errors during save
        logging.exception(f"Error saving uploaded file {filepath}: {e}")
        print(f"upload_file - Error saving file: {e}")
        print("upload_file - END (failure - save error)")
        return jsonify({"error": f"Error saving file: {e}"}), 500  # Use 500 for server error
    except Exception as e: #Catch any other error
        logging.exception(f"Error during file upload {filepath}: {e}")
        print(f"upload_file - Unexpected error: {e}")
        print("upload_file - END (failure - save error)")
        return jsonify({"error": f"Unexpected error during upload: {e}"}), 500


@app.route("/check_data", methods=["POST"])
def check_data():
    """Performs data quality checks on an uploaded file.

    This route receives the filename of an uploaded file via a POST request.
    It loads the data using `load_data` and performs data quality analysis
    using `analyze_data`. It returns a JSON response containing the analysis
    results.

    Returns:
        flask.Response: A JSON response with the analysis results and status code.
    """
    print("check_data - START")  # Debug print
    data = request.get_json()
    filename = data.get("filename")
    print(f"check_data - filename: {filename}")  # Debug print
    if not filename:
        print("check_data - Filename missing")  # Debug print
        return jsonify({"error": "Filename is required"}), 400
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    print(f"check_data - filepath: {filepath}")  # Debug print
    df = load_data(filepath, filename)
    if df is None:
        print("check_data - DataFrame is None")  # Debug print
        return jsonify({"error": "Error loading data"}), 500
    try:
        print("check_data - Before analyze_data")  # Debug print
        analysis_results = analyze_data(df)
        print(f"check_data - After analyze_data: results={analysis_results}")  # Debug print
        print("check_data - END (success)")  # Debug print
        return jsonify({"message": "Data quality check completed", "results": analysis_results}), 200
    except Exception as e:
        logging.error(f"Error during data quality check: {e}", exc_info=True)
        print(f"check_data - ERROR: {e}")  # Debug print
        print("check_data - END (error)")  # Debug print
        return jsonify({"error": "Error during check"}), 500

@app.route("/apply_fixes", methods=["POST"])
def apply_fixes():
    """Applies data fixes to an uploaded file.

    This route receives the filename of an uploaded file via a POST request.
    It loads the data, applies fixes using `apply_fixes_to_data`, and saves
    the corrected data back to the file.  It returns a JSON response
    indicating success or failure.

    Returns:
        flask.Response: A JSON response with a success/failure message and status code.
    """
    print("apply_fixes - START")  # Debug print
    data = request.get_json()
    print(f"apply_fixes - Data received: {data}")  # Debug print
    filename = data.get("filename")
    if not filename:
        print("apply_fixes - Filename missing")  # Debug print
        return jsonify({"error": "Filename is required"}), 400
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    print(f"apply_fixes - Filepath: {filepath}")  # Debug print
    df = load_data(filepath, filename)
    if df is None:
        print("apply_fixes - DataFrame is None")  # Debug print
        return jsonify({"error": "Error loading data"}), 500
    try:
        print("apply_fixes - Before apply_fixes_to_data")  # Debug print
        df_fixed, fixes_summary_str = apply_fixes_to_data(df)
        print(f"apply_fixes - After apply_fixes_to_data, df_fixed shape: {df_fixed.shape if df_fixed is not None else 'None'}")  # Debug print
        file_extension = filename.rsplit(".", 1)[1].lower()
        print(f"apply_fixes - File extension: {file_extension}")  # Debug print

        if file_extension == "csv":
            with open(filepath, 'w') as f:
                df_fixed.to_csv(f, index=False)
        elif file_extension in ("xls", "xlsx"):
            with pd.ExcelWriter(filepath) as writer:
                df_fixed.to_excel(writer, index=False)
        elif file_extension == "json":
            with open(filepath, 'w') as f:
                df_fixed.to_json(f, orient="records", lines=True)
        elif file_extension == "txt":
            with open(filepath, 'w') as f:
                df_fixed.to_csv(f, sep=",", index=False)
        elif file_extension == "tsv":
            with open(filepath, 'w') as f:
                df_fixed.to_csv(f, sep="\t", index=False)
        elif file_extension == "parquet":
            df_fixed.to_parquet(filepath)
        elif file_extension == "feather":
            df_fixed.to_feather(filepath)
        elif file_extension == "orc":
            df_fixed.to_orc(filepath)
        elif file_extension == "xml":
            df_fixed.to_xml(filepath)
        elif file_extension == "html":
            df_fixed.to_html(filepath)
        elif file_extension == "hdf5":
            with pd.HDFStore(filepath) as store:
                store.put("fixed_data", df_fixed)
        else:
            print(f"apply_fixes - Unsupported file format: {file_extension}")  # Debug print
            return jsonify({"error": "Unsupported file format for saving"}), 400
        print("apply_fixes - Fixes applied and saved")  # Debug print
        print("apply_fixes - END (success)")  # Debug print
        return jsonify({"message": "Fixes applied", "fixes_summary": fixes_summary_str}), 200
    except Exception as e:
        print(f"apply_fixes - Exception: {type(e).__name__}, {e}")  # Debug print
        traceback.print_exc()
        logging.error(f"Error applying fixes: {e}", exc_info=True)
        print("apply_fixes - END (error)")  # Debug print
        return jsonify({"error": "Error applying fixes"}), 500

#Modified Part of app.py

@app.route("/generate_plots", methods=["POST"])
def generate_plots():
    """Generates plot suggestions based on the uploaded data.

    This route receives the filename, selected model, and API key via a POST
    request.  It loads the data and uses the specified model ("gemini" or
    "claude") to generate plot suggestions.  It returns a JSON response
    containing the suggestions.

    Returns:
        flask.Response: A JSON response with plot suggestions and status code.
    """
    print("generate_plots - START") # Debug print
    data = request.get_json()
    filename, selected_model, api_key = data.get("filename"), data.get("selectedModel"), data.get("apiKey")
    print(f"generate_plots - filename: {filename}, selected_model: {selected_model}, api_key: {'PRESENT' if api_key else 'MISSING'}")  # Debug print
    if not filename or not api_key:
        logging.error("Filename and API key are required for plot generation")
        print("generate_plots - Filename or API key missing")  # Debug print
        return jsonify({"error": "Filename and API key are required"}), 400
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    print(f"generate_plots - filepath: {filepath}")  # Debug print
    df = load_data(filepath, filename)
    if df is None:
        print("generate_plots - DataFrame is None")  # Debug print
        return jsonify({"error": "Error loading data"}), 500

    try:
        if selected_model == "gemini":
            print("generate_plots - Calling get_plot_suggestion_from_gemini")  # Debug print
            plot_results = get_plot_suggestion_from_gemini(df, api_key)
        elif selected_model == "claude":
            print("generate_plots - Calling get_plot_suggestion_from_claude")  # Debug print
            plot_results = get_plot_suggestion_from_claude(df, api_key)
        else:
            logging.error(f"Invalid model selected for plot generation: {selected_model}")
            print(f"generate_plots - Invalid model: {selected_model}")  # Debug print
            return jsonify({"error": "Invalid model"}), 400

        if plot_results:
             return jsonify({"message": "Plots generated", "suggestions": plot_results}), 200
        else:
            return jsonify({"error": "Failed to generate plots"}), 500


    except Exception as e:
        logging.exception(f"Error in generate_plots: {e}")
        print(f"generate_plots - ERROR: {e}")  # Debug print
        print("generate_plots - END (error)") # Debug print

        #Even more descriptive error.
        return jsonify({"error": f"Error generating plots: {type(e).__name__}, {e}"}), 500
        

@app.route("/get_interpretation", methods=["POST"])
def get_interpretation():
    """Generates interpretations for a given plot suggestion.

    This route receives the selected model, API key, filename, and suggestion
    text via a POST request. It loads the data, creates a dataset summary,
    and uses the specified model ("gemini" or "claude") to generate an
    interpretation of the suggested plot.  It returns a JSON response
    containing the interpretation.

    Returns:
        flask.Response: A JSON response with the plot interpretation and status code.
    """
    print("get_interpretation - START")  # Debug print
    try:
        selected_model = request.json.get("selectedModel")
        api_key = request.json.get("apiKey")
        filename = request.json.get("filename")
        suggestion_text = request.json.get("suggestion")
        print(f"get_interpretation - selected_model: {selected_model}, api_key: {'PRESENT' if api_key else 'MISSING'}, filename: {filename}, suggestion_text: {suggestion_text}") # Debug print

        if not all([selected_model, api_key, filename, suggestion_text]):
            logging.error("Missing data for graph interpretation")
            print("get_interpretation - Missing data")  # Debug print
            return jsonify({"error": "Missing data"}), 400

        filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        print(f"get_interpretation - filepath: {filepath}")  # Debug print
        df = load_data(filepath, filename)
        if df is None:
            print("get_interpretation - DataFrame is None")  # Debug print
            return jsonify({"error": "Error loading data"}), 500

        print("get_interpretation - Before create_dataset_summary")  # Debug print
        dataset_summary = create_dataset_summary(df)
        print(f"get_interpretation - After create_dataset_summary: summary={dataset_summary[:100]}...")  # Debug print

        if selected_model == "gemini":
            print("get_interpretation - Calling generate_graph_interpretation_gemini")  # Debug print
            interpretation = generate_graph_interpretation_gemini(suggestion_text, dataset_summary, api_key)
            print(f"get_interpretation - generate_graph_interpretation_gemini returned: {interpretation}")  # Debug print
        elif selected_model == "claude":
            print("get_interpretation - Calling generate_graph_interpretation_claude")  # Debug print
            interpretation = generate_graph_interpretation_claude(suggestion_text, dataset_summary, api_key)
            print(f"get_interpretation - generate_graph_interpretation_claude returned: {interpretation}")  # Debug print
        else:
            print(f"get_interpretation - Invalid model: {selected_model}")  # Debug print
            return jsonify({"error": "Invalid model"}), 400

        if interpretation:
            print("get_interpretation - END (success)")  # Debug print
            return jsonify({"interpretation": interpretation}), 200
        else:
            logging.error(f"{selected_model} returned an empty interpretation")
            print(f"get_interpretation - Empty interpretation from {selected_model}")  # Debug print
            print("get_interpretation - END (failure)")  # Debug print
            return jsonify({"error": "Failed to generate interpretation"}), 500

    except Exception as e:
        logging.exception(f"Error in get_interpretation: {e}")
        print(f"get_interpretation - ERROR: {e}")  # Debug print
        print("get_interpretation - END (error)")  # Debug print
        return jsonify({"error": "Error generating interpretation"}), 500

@app.route("/graph_chat", methods=["POST"])
def graph_chat():
    """Handles user interactions with a graph image.

    This route receives the selected model, API key, user message, base64
    encoded image, and filename via a POST request.  It loads the data,
    creates a dataset summary, and uses the specified model ("gemini" or
    "claude") to handle communication related to the graph image. It returns a
    JSON response containing the model's response.

    Returns:
        flask.Response: A JSON response with the model's response and status code.
    """
    print("graph_chat - START")  # Debug print
    try:
        selected_model, api_key, user_message, base64_image, filename = [request.json.get(key) for key in ("selectedModel", "apiKey", "message", "image", "filename")]
        print(f"graph_chat - selected_model: {selected_model}, api_key: {'PRESENT' if api_key else 'MISSING'}, user_message: {user_message}, base64_image: {'PRESENT' if base64_image else 'MISSING'}, filename: {filename}") # Debug print


        if not all([selected_model, api_key, user_message, base64_image, filename]):
            logging.error("Missing data for graph chat")
            print("graph_chat - Missing data")  # Debug print
            return jsonify({"error": "Missing data"}), 400

        image_data = base64.b64decode(base64_image)
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        print(f"graph_chat - filepath: {filepath}")  # Debug print
        df = load_data(filepath, filename)
        if df is None:
            print("graph_chat - DataFrame is None") # Debug print
            return jsonify({"error": "Error loading data"}), 500

        print("graph_chat - Before create_dataset_summary")  # Debug print
        dataset_summary = create_dataset_summary(df)
        print(f"graph_chat - After create_dataset_summary: summary={dataset_summary[:100]}...")  # Debug print

        if selected_model == "gemini":
            print("graph_chat - Calling handle_graph_communication_gemini")  # Debug print
            response = handle_graph_communication_gemini(image_data, dataset_summary, user_message, api_key)
            print(f"graph_chat - handle_graph_communication_gemini returned: {response}")  # Debug print
        elif selected_model == "claude":
            print("graph_chat - Calling handle_graph_communication_claude")  # Debug print
            response = handle_graph_communication_claude(image_data, dataset_summary, user_message, api_key)
            print(f"graph_chat - handle_graph_communication_claude returned: {response}")  # Debug print
        else:
            print(f"graph_chat - Invalid model: {selected_model}")  # Debug print
            return jsonify({"error": "Invalid model"}), 400


        if response:
            print("graph_chat - END (success)")  # Debug print
            return jsonify({"response": response}), 200
        else:
            logging.error(f"{selected_model} returned an empty response in graph chat")
            print(f"graph_chat - Empty response from {selected_model}")  # Debug print
            print("graph_chat - END (failure)")  # Debug print
            return jsonify({"error": f"Failed to get response from {selected_model}"}), 500

    except Exception as e:
        logging.exception(f"Error in graph_chat: {e}")
        print(f"graph_chat - ERROR: {e}")  # Debug print
        print("graph_chat - END (error)")  # Debug print
        return jsonify({"error": "Error in chat"}), 500

@app.route("/uploads/<filename>")
def uploaded_file(filename):
    """Serves uploaded files.

    This route allows direct access to uploaded files via their filenames.

    Args:
        filename (str): The name of the file to retrieve.

    Returns:
        flask.Response: The requested file, served with the correct MIME type.
    """
    print(f"uploaded_file - START: filename={filename}")  # Debug print
    print(f"uploaded_file - END: filename={filename}")  # Debug print
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)

if __name__ == "__main__":
    app.run(debug=True)